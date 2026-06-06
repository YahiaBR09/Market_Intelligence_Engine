import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# فتح ملف الإكسيل الذي أنتجه الـ Matcher
wb = openpyxl.load_workbook("Smart_Competitor_Analysis.xlsx")
ws = wb.active

# 1. تنسيق السطر الأول (الرأس) باللون الكحلي الفخم والكتابة البيضاء العريضة
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment

# 2. تلوين حالات المقارنة (Status) تلقائياً بناءً على النتيجة لتسهيل القراءة بالعين
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # أرخص من المنافس
red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # أغلى من المنافس
green_font = Font(name="Segoe UI", size=10, color="375623", bold=True)
red_font = Font(name="Segoe UI", size=10, color="C65911", bold=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# المرور على الصفوف لتنسيقها وتوسيع الأعمدة تلقائياً
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        cell.font = Font(name="Segoe UI", size=10)
        
        # تلوين عمود الحالة (Status) وهو العمود رقم 7 غالباً
        if cell.column == 7: 
            if cell.value == "We are Cheaper":
                cell.fill = green_fill
                cell.font = green_font
            elif cell.value == "Competitor is Cheaper":
                cell.fill = red_fill
                cell.font = red_font
            cell.alignment = center_alignment

# ضبط أبعاد الأعمدة تلقائياً لتناسب حجم النصوص بدون ظهور رموز (###)
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# حفظ التعديلات الجمالية الفاخرة
wb.save("Smart_Competitor_Analysis.xlsx")
print("✨ تم تلوين وتنسيق تقرير الإكسيل بشكل استراتيجي باهر جاهز للعرض المباشر!")