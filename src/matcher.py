import csv
import re
import pandas as pd
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ملفات المدخلات والمخرجات
customer_file = "./data/customer_real_prices.csv"
competitor_file = "./data/competitor_real_prices.csv"
output_file = "./data/Smart_Competitor_Analysis.xlsx"

print("🧠 جاري تشغيل خوارزمية المطابقة الذكية والصارمة ومنع خلط الأصناف...")

def clean_text(text):
    if not text or pd.isna(text):
        return ""
    text = str(text).lower()
    text = re.sub(r'[\-\_\,\.\/\(\)]', ' ', text)
    return " ".join(text.split())

def get_core_tool_type(cleaned_title):
    """تحديد الفئة الدقيقة لمنع مطابقة Screwdriver مع Extractor"""
    if "extractor" in cleaned_title:
        return "extractor"
    elif "screwdriver" in cleaned_title:
        return "screwdriver"
    elif "spanner" in cleaned_title or "wrench" in cleaned_title:
        return "spanner"
    elif "socket" in cleaned_title:
        return "socket"
    elif "plier" in cleaned_title:
        return "pliers"
    return "other"

def extract_numbers(text):
    return set(re.findall(r'\d+', text))

# قراءة البيانات
try:
    df_customer = pd.read_csv(customer_file)
    df_competitor = pd.read_csv(competitor_file)
except Exception as e:
    print(f"🚨 خطأ في قراءة ملفات الـ CSV: {e}")
    exit()

matched_data = []

# حلقة المطابقة الذكية
for idx_cust, row_cust in df_customer.iterrows():
    cust_title = row_cust['Our Product Name']
    cust_price = float(row_cust['Our Price']) if pd.notna(row_cust['Our Price']) else 0.0
    cust_sku = row_cust['Our SKU']
    
    cust_clean = clean_text(cust_title)
    cust_core = get_core_tool_type(cust_clean)
    cust_nums = extract_numbers(cust_clean)
    
    best_match_row = None
    best_score = 0.0
    
    for idx_comp, row_comp in df_competitor.iterrows():
        comp_title = row_comp['Product Name']
        comp_clean = clean_text(comp_title)
        
        # شرط الفئة الصارم
        comp_core = get_core_tool_type(comp_clean)
        if cust_core != comp_core:
            continue
            
        # شرط المقاسات الرقمية
        comp_nums = extract_numbers(comp_clean)
        if cust_nums and comp_nums and not (cust_nums & comp_nums):
            continue
            
        score = SequenceMatcher(None, cust_clean, comp_clean).ratio()
        if score > best_score:
            best_score = score
            best_match_row = row_comp

    if best_match_row is not None and best_score >= 0.50:
        comp_title = best_match_row['Product Name']
        comp_price = float(best_match_row['Price (FJD)']) if pd.notna(best_match_row['Price (FJD)']) else 0.0
        comp_url = best_match_row['Product URL']
        comp_sku = best_match_row['SKU/ID']
        
        price_diff = round(cust_price - comp_price, 2)
        status = "We are Cheaper" if price_diff < 0 else ("Competitor is Cheaper" if price_diff > 0 else "Equal Price")
        
        matched_data.append([
            cust_sku, cust_title, cust_price,
            comp_sku, comp_title, comp_price,
            status, price_diff, comp_url
        ])
    else:
        matched_data.append([
            cust_sku, cust_title, cust_price,
            "N/A", "No Direct Match Found", 0.0,
            "Unique Product", 0.0, "N/A"
        ])

# حفظ البيانات في ملف إكسيل مبدئي
columns = [
    'Our SKU', 'Our Product Name', 'Our Price (FJD)',
    'Competitor SKU', 'Competitor Product Name', 'Competitor Price (FJD)',
    'Status', 'Price Difference', 'Competitor URL'
]
df_result = pd.DataFrame(matched_data, columns=columns)
df_result.to_excel(output_file, index=False)

# 🎨 ------------------- قسم التلوين والتنسيق الجمالي التلقائي -------------------
print("🎨 جاري تطبيق التنسيق الجمالي الفخم وتلوين الحالات لراحة عين العميل...")
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# تنسيق الرأس (كحلي ملكي عريض)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment

# تنسيقات الصفوف والحالات
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # نحن أرخص
red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # المنافس أرخص
blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # منتج فريد لنا

green_font = Font(name="Segoe UI", size=10, color="375623", bold=True)
red_font = Font(name="Segoe UI", size=10, color="C65911", bold=True)
blue_font = Font(name="Segoe UI", size=10, color="1F4E78", bold=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        cell.font = Font(name="Segoe UI", size=10)
        
        # تلوين وتنسيق عمود الحالة (العمود السابع)
        if cell.column == 7:
            if cell.value == "We are Cheaper":
                cell.fill = green_fill
                cell.font = green_font
            elif cell.value == "Competitor is Cheaper":
                cell.fill = red_fill
                cell.font = red_font
            elif cell.value == "Unique Product":
                cell.fill = blue_fill
                cell.font = blue_font
            cell.alignment = center_alignment

# ضبط أبعاد الأعمدة تلقائياً لتناسب النصوص بالكامل
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(output_file)
print(f"✨ تم إنتاج لوحة التحكم وتنسيق الملف بنجاح باهر واكتمل المشروع بالكامل! 🏆")
